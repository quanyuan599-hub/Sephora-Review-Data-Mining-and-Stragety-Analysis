from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import json

import numpy as np
import pandas as pd
from PIL import Image, ImageDraw, ImageFont


EXPECTED_COLUMNS = [
    "Unnamed: 0",
    "author_id",
    "rating",
    "is_recommended",
    "helpfulness",
    "total_feedback_count",
    "total_neg_feedback_count",
    "total_pos_feedback_count",
    "submission_time",
    "review_text",
    "review_title",
    "skin_tone",
    "eye_color",
    "skin_type",
    "hair_color",
    "product_id",
    "product_name",
    "brand_name",
    "price_usd",
]

TARGET_BRANDS = ["Drunk Elephant", "Tatcha", "The Ordinary"]
BRAND_KEY = {brand.casefold(): brand for brand in TARGET_BRANDS}
KEY_MISSINGNESS_COLUMNS = [
    "review_text",
    "is_recommended",
    "rating",
    "submission_time",
    "submission_date",
    "price_usd",
    "price_usd_clean",
    "brand_name",
    "product_id",
    "product_name",
    "skin_type",
    "skin_tone",
    "eye_color",
    "hair_color",
    "review_title",
    "helpfulness",
]


@dataclass(frozen=True)
class AnalysisConfig:
    input_dir: Path
    output_dir: Path
    sample_per_brand: int = 2000
    top_products: int = 5
    seed: int = 214


def run_analysis(config: AnalysisConfig) -> dict[str, object]:
    tables_dir = config.output_dir / "tables"
    figures_dir = config.output_dir / "figures"
    tables_dir.mkdir(parents=True, exist_ok=True)
    figures_dir.mkdir(parents=True, exist_ok=True)

    raw, source_log = read_raw_files(config.input_dir)
    raw_rows = len(raw)
    missing_raw = missingness_summary(raw, "raw_combined")

    cleaned, cleaning_log = clean_reviews(raw)
    cleaned_rows = len(cleaned)

    brand_filtered = cleaned[cleaned["brand_name_norm"].isin(TARGET_BRANDS)].copy()
    top_ids = select_top_product_ids(brand_filtered, config.top_products)
    top_filtered = brand_filtered[brand_filtered["brand_product_key"].isin(top_ids)].copy()
    sampled = sample_by_brand(top_filtered, config.sample_per_brand, config.seed)

    table_map = build_tables(raw, cleaned, brand_filtered, top_filtered, sampled, source_log, cleaning_log)
    write_tables(table_map, tables_dir)
    write_excel(table_map, config.output_dir / "sephora_review_analysis_outputs.xlsx")
    sampled.to_csv(config.output_dir / "cleaned_sample.csv", index=False, encoding="utf-8-sig")
    write_final_dataset_workbook(sampled, config.output_dir / "sephora_target_brand_6000_reviews.xlsx")
    write_analysis_summary_workbook(table_map, config.output_dir / "sephora_target_brand_analysis_summary.xlsx")

    write_figures(table_map, figures_dir)
    brief_path = write_brief(
        config.output_dir / "task1_brief.md",
        raw_rows,
        cleaned_rows,
        brand_filtered,
        top_filtered,
        sampled,
        table_map,
    )

    metadata = {
        "raw_rows": raw_rows,
        "cleaned_rows": cleaned_rows,
        "target_brand_rows_before_top_products": len(brand_filtered),
        "top_product_rows_before_sampling": len(top_filtered),
        "final_rows": len(sampled),
        "brief_path": str(brief_path),
    }
    (config.output_dir / "run_metadata.json").write_text(
        json.dumps(metadata, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    return metadata


def read_raw_files(input_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    files = sorted(input_dir.glob("reviews*.csv"))
    if not files:
        raise FileNotFoundError(f"No reviews*.csv files found in {input_dir}")

    frames: list[pd.DataFrame] = []
    source_rows: list[dict[str, object]] = []
    for file in files:
        df = pd.read_csv(
            file,
            usecols=EXPECTED_COLUMNS,
            encoding="latin1",
            low_memory=False,
        )
        df["source_file"] = file.name
        frames.append(df)
        source_rows.append({"source_file": file.name, "raw_rows": len(df)})

    return pd.concat(frames, ignore_index=True), pd.DataFrame(source_rows)


def clean_reviews(raw: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = raw.copy()
    start_rows = len(df)
    log_rows: list[dict[str, object]] = [{"step": "raw_rows", "rows": start_rows, "removed": 0}]

    for col in ["review_text", "review_title", "skin_type", "product_id", "product_name", "brand_name"]:
        df[col] = df[col].astype("string").str.strip()

    df = df.drop_duplicates()
    log_rows.append({"step": "drop_exact_duplicate_rows", "rows": len(df), "removed": start_rows - len(df)})
    before = len(df)
    key_cols = ["author_id", "product_id", "submission_time", "review_text"]
    df = df.drop_duplicates(subset=key_cols, keep="first")
    log_rows.append({"step": "drop_duplicate_review_keys", "rows": len(df), "removed": before - len(df)})

    before = len(df)
    df = df[df["review_text"].notna() & df["review_text"].str.len().gt(0)].copy()
    log_rows.append({"step": "drop_missing_review_text", "rows": len(df), "removed": before - len(df)})

    before = len(df)
    df["is_recommended"] = pd.to_numeric(df["is_recommended"], errors="coerce")
    df = df[df["is_recommended"].isin([0, 1])].copy()
    df["is_recommended"] = df["is_recommended"].astype(int)
    log_rows.append({"step": "drop_missing_or_invalid_is_recommended", "rows": len(df), "removed": before - len(df)})

    before = len(df)
    df["rating"] = pd.to_numeric(df["rating"], errors="coerce")
    df = df[df["rating"].between(1, 5, inclusive="both")].copy()
    df["rating"] = df["rating"].astype(int)
    log_rows.append({"step": "drop_rating_outside_1_to_5", "rows": len(df), "removed": before - len(df)})

    before_missing_date = df["submission_time"].isna().sum()
    df["submission_date"] = pd.to_datetime(df["submission_time"], errors="coerce", format="mixed")
    invalid_dates = int(df["submission_date"].isna().sum() - before_missing_date)
    df["month"] = df["submission_date"].dt.to_period("M").astype("string")
    df["quarter"] = df["submission_date"].dt.to_period("Q").astype("string")
    log_rows.append({"step": "parse_submission_time", "rows": len(df), "removed": 0, "invalid_dates": invalid_dates})

    price_before = df["price_usd"].isna().sum()
    df["price_usd_original"] = df["price_usd"]
    df["price_usd"] = pd.to_numeric(df["price_usd"], errors="coerce")
    coerced_missing = int(df["price_usd"].isna().sum() - price_before)
    nonpositive = int(df["price_usd"].le(0).fillna(False).sum())
    df.loc[df["price_usd"].le(0), "price_usd"] = np.nan
    q1, q3 = df["price_usd"].quantile([0.25, 0.75])
    iqr = q3 - q1
    upper = q3 + 3 * iqr
    outlier_mask = df["price_usd"].gt(upper)
    price_outliers = int(outlier_mask.sum())
    df.loc[outlier_mask, "price_usd"] = np.nan
    df["price_usd_clean"] = df["price_usd"]
    df["price_usd_clean"] = df["price_usd_clean"].fillna(
        df.groupby("product_id")["price_usd_clean"].transform("median")
    )
    df["price_usd_clean"] = df["price_usd_clean"].fillna(
        df.groupby("brand_name")["price_usd_clean"].transform("median")
    )
    df["price_usd_clean"] = df["price_usd_clean"].fillna(df["price_usd_clean"].median())
    log_rows.append(
        {
            "step": "clean_price_usd",
            "rows": len(df),
            "removed": 0,
            "coerced_to_missing": coerced_missing,
            "nonpositive_set_missing": nonpositive,
            "high_outliers_set_missing": price_outliers,
            "high_outlier_threshold": round(float(upper), 2),
        }
    )

    df["brand_name_norm"] = df["brand_name"].str.casefold().map(BRAND_KEY).fillna(df["brand_name"])
    df["skin_type"] = df["skin_type"].replace({"": pd.NA}).fillna("Unknown")
    df["product_name"] = df["product_name"].replace({"": pd.NA}).fillna("Unknown product")
    df["brand_product_key"] = df["brand_name_norm"].astype(str) + "||" + df["product_id"].astype(str)
    return df, pd.DataFrame(log_rows)


def missingness_summary(df: pd.DataFrame, dataset: str) -> pd.DataFrame:
    return (
        pd.DataFrame(
            {
                "dataset": dataset,
                "column": df.columns,
                "missing_count": [int(df[col].isna().sum()) for col in df.columns],
                "missing_pct": [round(float(df[col].isna().mean() * 100), 2) for col in df.columns],
            }
        )
        .sort_values(["missing_pct", "column"], ascending=[False, True])
        .reset_index(drop=True)
    )


def missingness_overview(datasets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for name, df in datasets.items():
        total_cells = int(df.shape[0] * df.shape[1])
        total_missing = int(df.isna().sum().sum())
        critical_cols = [col for col in ["review_text", "is_recommended", "rating", "submission_time", "price_usd"] if col in df]
        critical_missing = int(df[critical_cols].isna().sum().sum()) if critical_cols else 0
        rows.append(
            {
                "dataset": name,
                "rows": len(df),
                "columns_checked": df.shape[1],
                "total_missing_cells": total_missing,
                "overall_missing_pct": round(total_missing / total_cells * 100, 2) if total_cells else 0,
                "columns_with_missing": int(df.isna().any().sum()),
                "critical_fields_missing_cells": critical_missing,
            }
        )
    return pd.DataFrame(rows)


def key_missingness_summary(datasets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for dataset, df in datasets.items():
        for col in KEY_MISSINGNESS_COLUMNS:
            if col not in df.columns:
                continue
            missing = int(df[col].isna().sum())
            rows.append(
                {
                    "dataset": dataset,
                    "field_group": classify_missingness_field(col),
                    "column": col,
                    "missing_count": missing,
                    "missing_pct": round(float(df[col].isna().mean() * 100), 2),
                    "included_reason": missingness_reason(col),
                }
            )
    return (
        pd.DataFrame(rows)
        .sort_values(["dataset", "field_group", "missing_pct", "column"], ascending=[True, True, False, True])
        .reset_index(drop=True)
    )


def classify_missingness_field(col: str) -> str:
    if col in {"review_text", "is_recommended", "rating", "submission_time", "submission_date"}:
        return "Required cleaning fields"
    if col in {"brand_name", "product_id", "product_name", "price_usd", "price_usd_clean"}:
        return "Brand/product/price fields"
    if col in {"skin_type", "skin_tone", "eye_color", "hair_color"}:
        return "Customer profile fields"
    return "Optional review metadata"


def missingness_reason(col: str) -> str:
    reasons = {
        "review_text": "required; rows with missing text are removed",
        "is_recommended": "required KPI field; rows with missing/invalid values are removed",
        "rating": "required quality check; must be 1-5",
        "submission_time": "required for trend analysis",
        "submission_date": "parsed date used for monthly/quarterly trend",
        "price_usd": "cleaned for price outlier handling",
        "price_usd_clean": "imputed cleaned price used in KPI tables",
        "brand_name": "required for brand filtering",
        "product_id": "required for top product selection",
        "product_name": "used for product labels",
        "skin_type": "used for skin type KPI",
        "skin_tone": "optional customer profile field",
        "eye_color": "optional customer profile field",
        "hair_color": "optional customer profile field",
        "review_title": "optional review metadata",
        "helpfulness": "optional engagement metadata",
    }
    return reasons.get(col, "selected analysis field")


def select_top_product_ids(df: pd.DataFrame, top_n: int) -> set[str]:
    volume = (
        df.groupby(["brand_name_norm", "product_id", "product_name"], dropna=False)
        .size()
        .reset_index(name="review_count")
        .sort_values(["brand_name_norm", "review_count", "product_name"], ascending=[True, False, True])
    )
    return set(
        volume.groupby("brand_name_norm")
        .head(top_n)
        .assign(brand_product_key=lambda x: x["brand_name_norm"].astype(str) + "||" + x["product_id"].astype(str))[
            "brand_product_key"
        ]
    )


def sample_by_brand(df: pd.DataFrame, sample_per_brand: int, seed: int) -> pd.DataFrame:
    parts = []
    for _, group in df.groupby("brand_name_norm", sort=True):
        n = min(sample_per_brand, len(group))
        parts.append(group.sample(n=n, random_state=seed) if len(group) > n else group)
    return pd.concat(parts, ignore_index=True).sort_values(["brand_name_norm", "submission_date"]).reset_index(drop=True)


def build_tables(
    raw: pd.DataFrame,
    cleaned: pd.DataFrame,
    brand_filtered: pd.DataFrame,
    top_filtered: pd.DataFrame,
    sampled: pd.DataFrame,
    source_log: pd.DataFrame,
    cleaning_log: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    datasets = {
        "raw_combined": raw,
        "cleaned_all_brands": cleaned,
        "sampled_target_brands_top_products": sampled,
    }
    raw_missing = missingness_summary(raw, "raw_combined")
    clean_missing = missingness_summary(cleaned, "cleaned_all_brands")
    sample_missing = missingness_summary(sampled, "sampled_target_brands_top_products")
    missingness_detail = pd.concat([raw_missing, clean_missing, sample_missing], ignore_index=True)
    missingness = key_missingness_summary(datasets)
    missingness_stage_overview = missingness_overview(datasets)

    brand_kpi = aggregate_kpis(sampled, ["brand_name_norm"]).rename(columns={"brand_name_norm": "brand"})
    product_volume = aggregate_kpis(sampled, ["brand_name_norm", "product_id", "product_name"]).rename(
        columns={"brand_name_norm": "brand"}
    )
    product_volume = product_volume.sort_values(["brand", "review_count"], ascending=[True, False])
    skin_type_kpi = aggregate_kpis(sampled, ["brand_name_norm", "skin_type"]).rename(columns={"brand_name_norm": "brand"})
    monthly_trend = aggregate_kpis(sampled[sampled["month"].notna()], ["brand_name_norm", "month"]).rename(
        columns={"brand_name_norm": "brand"}
    )
    quarterly_trend = aggregate_kpis(sampled[sampled["quarter"].notna()], ["brand_name_norm", "quarter"]).rename(
        columns={"brand_name_norm": "brand"}
    )
    bias_uncertainty = build_bias_uncertainty(sampled)
    consistency_checks = build_consistency_checks(sampled, brand_kpi, product_volume, skin_type_kpi, quarterly_trend)

    distribution = {
        "source_file_summary": source_log,
        "cleaning_log": cleaning_log,
        "missingness_overview": missingness_stage_overview,
        "missingness_summary": missingness,
        "missingness_detail_full": missingness_detail,
        "brand_kpi": brand_kpi,
        "product_volume": product_volume,
        "skin_type_kpi": skin_type_kpi,
        "monthly_trend": monthly_trend,
        "quarterly_trend": quarterly_trend,
        "bias_uncertainty": bias_uncertainty,
        "consistency_checks": consistency_checks,
        "target_brand_rows_before_sampling": aggregate_kpis(brand_filtered, ["brand_name_norm"]).rename(
            columns={"brand_name_norm": "brand"}
        ),
        "top_product_rows_before_sampling": aggregate_kpis(top_filtered, ["brand_name_norm"]).rename(
            columns={"brand_name_norm": "brand"}
        ),
    }
    return distribution


def build_bias_uncertainty(sampled: pd.DataFrame) -> pd.DataFrame:
    quantitative_parts = [
        uncertainty_for_groups(sampled, ["brand_name_norm"], "brand").rename(columns={"brand_name_norm": "brand"}),
        uncertainty_for_groups(sampled, ["brand_name_norm", "skin_type"], "brand_skin_type").rename(
            columns={"brand_name_norm": "brand"}
        ),
        uncertainty_for_groups(sampled[sampled["quarter"].notna()], ["brand_name_norm", "quarter"], "brand_quarter").rename(
            columns={"brand_name_norm": "brand"}
        ),
    ]
    quantitative = pd.concat(quantitative_parts, ignore_index=True)
    quantitative["bias_or_uncertainty_type"] = "sampling_uncertainty"
    quantitative["note"] = quantitative["review_count"].apply(sample_size_note)

    notes = pd.DataFrame(
        [
            {
                "level": "dataset",
                "brand": "All target brands",
                "segment": "analysis scope",
                "review_count": len(sampled),
                "avg_rating": np.nan,
                "avg_rating_ci95_low": np.nan,
                "avg_rating_ci95_high": np.nan,
                "recommendation_rate": np.nan,
                "recommendation_rate_ci95_low": np.nan,
                "recommendation_rate_ci95_high": np.nan,
                "bias_or_uncertainty_type": "selection_bias",
                "note": "Final dataset only includes 3 target brands, each brand's top 5 products, and up to 2,000 sampled reviews per brand; it should not be generalized to all Sephora reviews.",
            },
            {
                "level": "dataset",
                "brand": "All target brands",
                "segment": "required-field filtering",
                "review_count": len(sampled),
                "avg_rating": np.nan,
                "avg_rating_ci95_low": np.nan,
                "avg_rating_ci95_high": np.nan,
                "recommendation_rate": np.nan,
                "recommendation_rate_ci95_low": np.nan,
                "recommendation_rate_ci95_high": np.nan,
                "bias_or_uncertainty_type": "missingness_bias",
                "note": "Rows missing review_text or is_recommended were removed before brand/product sampling, so results describe reviews with complete required fields.",
            },
            {
                "level": "dataset",
                "brand": "All target brands",
                "segment": "time trend",
                "review_count": len(sampled),
                "avg_rating": np.nan,
                "avg_rating_ci95_low": np.nan,
                "avg_rating_ci95_high": np.nan,
                "recommendation_rate": np.nan,
                "recommendation_rate_ci95_low": np.nan,
                "recommendation_rate_ci95_high": np.nan,
                "bias_or_uncertainty_type": "low_volume_periods",
                "note": "Quarterly estimates are more stable than monthly estimates, but early/low-volume quarters should still be interpreted cautiously.",
            },
        ]
    )
    columns = [
        "level",
        "brand",
        "segment",
        "review_count",
        "avg_rating",
        "avg_rating_ci95_low",
        "avg_rating_ci95_high",
        "recommendation_rate",
        "recommendation_rate_ci95_low",
        "recommendation_rate_ci95_high",
        "bias_or_uncertainty_type",
        "note",
    ]
    return pd.concat([notes[columns], quantitative[columns]], ignore_index=True)


def uncertainty_for_groups(df: pd.DataFrame, group_cols: list[str], level: str) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    if df.empty:
        return pd.DataFrame()
    for keys, group in df.groupby(group_cols, dropna=False):
        if not isinstance(keys, tuple):
            keys = (keys,)
        n = len(group)
        rating_mean = float(group["rating"].mean())
        rating_sd = float(group["rating"].std(ddof=1)) if n > 1 else 0.0
        rating_margin = 1.96 * rating_sd / np.sqrt(n) if n > 1 else 0.0
        rec_rate = float(group["is_recommended"].mean() * 100)
        p = rec_rate / 100
        rec_margin = 1.96 * np.sqrt(p * (1 - p) / n) * 100 if n > 0 else np.nan
        row = {
            "level": level,
            "review_count": n,
            "avg_rating": round(rating_mean, 2),
            "avg_rating_ci95_low": round(max(1.0, rating_mean - rating_margin), 2),
            "avg_rating_ci95_high": round(min(5.0, rating_mean + rating_margin), 2),
            "recommendation_rate": round(rec_rate, 1),
            "recommendation_rate_ci95_low": round(max(0.0, rec_rate - rec_margin), 1),
            "recommendation_rate_ci95_high": round(min(100.0, rec_rate + rec_margin), 1),
        }
        for col, value in zip(group_cols, keys):
            row[col] = value
        rows.append(row)
    out = pd.DataFrame(rows)
    if len(group_cols) == 1:
        out["segment"] = "all selected products"
    elif len(group_cols) >= 2:
        out["segment"] = out[group_cols[-1]].astype(str)
    return out


def sample_size_note(n: int) -> str:
    if n < 30:
        return "Very small sample; use directionally only."
    if n < 100:
        return "Small sample; confidence interval is relatively wide."
    return "Adequate sample size for descriptive comparison within this selected dataset."


def build_consistency_checks(
    sampled: pd.DataFrame,
    brand_kpi: pd.DataFrame,
    product_volume: pd.DataFrame,
    skin_type_kpi: pd.DataFrame,
    quarterly_trend: pd.DataFrame,
) -> pd.DataFrame:
    checks = [
        {
            "check": "final_row_count",
            "expected": 6000,
            "actual": len(sampled),
            "status": "PASS" if len(sampled) == 6000 else "CHECK",
            "detail": "Complete final dataset should contain 2,000 rows per target brand.",
        },
        {
            "check": "brand_count",
            "expected": 3,
            "actual": sampled["brand_name_norm"].nunique(),
            "status": "PASS" if sampled["brand_name_norm"].nunique() == 3 else "CHECK",
            "detail": "Only Drunk Elephant, Tatcha, and The Ordinary are included.",
        },
        {
            "check": "per_brand_rows",
            "expected": "2,000 each",
            "actual": "; ".join(f"{k}: {v}" for k, v in sampled["brand_name_norm"].value_counts().sort_index().items()),
            "status": "PASS" if sampled["brand_name_norm"].value_counts().eq(2000).all() else "CHECK",
            "detail": "Sampling cap is applied independently by brand.",
        },
        {
            "check": "top_product_count",
            "expected": 15,
            "actual": product_volume[["brand", "product_id"]].drop_duplicates().shape[0],
            "status": "PASS" if product_volume[["brand", "product_id"]].drop_duplicates().shape[0] == 15 else "CHECK",
            "detail": "Each brand should contribute 5 products.",
        },
        {
            "check": "product_counts_sum_to_final_rows",
            "expected": len(sampled),
            "actual": int(product_volume["review_count"].sum()),
            "status": "PASS" if int(product_volume["review_count"].sum()) == len(sampled) else "CHECK",
            "detail": "Product distribution is calculated from the same 6,000-row final dataset.",
        },
        {
            "check": "brand_counts_sum_to_final_rows",
            "expected": len(sampled),
            "actual": int(brand_kpi["review_count"].sum()),
            "status": "PASS" if int(brand_kpi["review_count"].sum()) == len(sampled) else "CHECK",
            "detail": "Brand KPI is calculated from the same 6,000-row final dataset.",
        },
        {
            "check": "skin_type_counts_sum_to_final_rows",
            "expected": len(sampled),
            "actual": int(skin_type_kpi["review_count"].sum()),
            "status": "PASS" if int(skin_type_kpi["review_count"].sum()) == len(sampled) else "CHECK",
            "detail": "Missing skin_type is converted to Unknown, so grouped rows should sum to final rows.",
        },
        {
            "check": "quarterly_counts_sum_to_final_rows",
            "expected": len(sampled),
            "actual": int(quarterly_trend["review_count"].sum()),
            "status": "PASS" if int(quarterly_trend["review_count"].sum()) == len(sampled) else "CHECK",
            "detail": "All final rows have valid parsed submission_date and quarter.",
        },
        {
            "check": "required_fields_no_missing",
            "expected": 0,
            "actual": int(sampled[["review_text", "is_recommended", "rating", "submission_date"]].isna().sum().sum()),
            "status": "PASS"
            if int(sampled[["review_text", "is_recommended", "rating", "submission_date"]].isna().sum().sum()) == 0
            else "CHECK",
            "detail": "Required analysis fields should have no missing values in the final dataset.",
        },
    ]
    return pd.DataFrame(checks)


def aggregate_kpis(df: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=group_cols + ["review_count", "avg_rating", "recommendation_rate", "avg_price_usd"])
    grouped = df.groupby(group_cols, dropna=False)
    out = grouped.agg(
        review_count=("review_text", "size"),
        avg_rating=("rating", "mean"),
        recommendation_rate=("is_recommended", "mean"),
        avg_price_usd=("price_usd_clean", "mean"),
        unique_products=("product_id", "nunique"),
        date_min=("submission_date", "min"),
        date_max=("submission_date", "max"),
    ).reset_index()
    out["avg_rating"] = out["avg_rating"].round(2)
    out["recommendation_rate"] = (out["recommendation_rate"] * 100).round(1)
    out["avg_price_usd"] = out["avg_price_usd"].round(2)
    out["date_min"] = out["date_min"].dt.strftime("%Y-%m-%d")
    out["date_max"] = out["date_max"].dt.strftime("%Y-%m-%d")
    return out


def write_tables(table_map: dict[str, pd.DataFrame], tables_dir: Path) -> None:
    for name, df in table_map.items():
        df.to_csv(tables_dir / f"{name}.csv", index=False, encoding="utf-8-sig")


def write_excel(table_map: dict[str, pd.DataFrame], path: Path) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in table_map.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
    format_workbook(path)


def write_final_dataset_workbook(sampled: pd.DataFrame, path: Path) -> None:
    ordered_cols = [
        "brand_name_norm",
        "product_id",
        "product_name",
        "rating",
        "is_recommended",
        "submission_date",
        "month",
        "quarter",
        "price_usd_clean",
        "review_text",
        "review_title",
        "skin_type",
        "skin_tone",
        "eye_color",
        "hair_color",
        "author_id",
        "helpfulness",
        "total_feedback_count",
        "total_neg_feedback_count",
        "total_pos_feedback_count",
        "source_file",
    ]
    cols = [col for col in ordered_cols if col in sampled.columns]
    df = sampled[cols].copy()
    df = df.rename(columns={"brand_name_norm": "brand"})
    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        df.to_excel(writer, sheet_name="final_6000_reviews", index=False)
    format_workbook(path, freeze_panes="A2")


def write_analysis_summary_workbook(table_map: dict[str, pd.DataFrame], path: Path) -> None:
    sheets = {
        "missingness_summary": table_map["missingness_summary"],
        "brand_rating_recommendation": table_map["brand_kpi"][
            ["brand", "review_count", "recommendation_rate", "avg_rating", "avg_price_usd", "date_min", "date_max"]
        ],
        "top5_product_review_count": table_map["product_volume"][
            ["brand", "product_id", "product_name", "review_count", "recommendation_rate", "avg_rating"]
        ],
        "skin_type_kpi": table_map["skin_type_kpi"][
            ["brand", "skin_type", "review_count", "recommendation_rate", "avg_rating"]
        ],
        "quarterly_trend": table_map["quarterly_trend"][
            ["brand", "quarter", "review_count", "recommendation_rate", "avg_rating"]
        ],
        "bias_uncertainty": table_map["bias_uncertainty"],
        "consistency_checks": table_map["consistency_checks"],
    }
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
    format_workbook(path)


def format_workbook(path: Path, freeze_panes: str = "A2") -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = freeze_panes
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column_cells[:200]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 55)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def write_figures(table_map: dict[str, pd.DataFrame], figures_dir: Path) -> None:
    brand_kpi = table_map["brand_kpi"]
    product_volume = table_map["product_volume"]
    skin_type_kpi = table_map["skin_type_kpi"]
    monthly_trend = table_map["monthly_trend"]
    quarterly_trend = table_map["quarterly_trend"]

    draw_bar_chart(
        brand_kpi["brand"].tolist(),
        brand_kpi["review_count"].tolist(),
        "Brand review volume",
        figures_dir / "brand_review_volume.png",
    )
    draw_bar_chart(
        brand_kpi["brand"].tolist(),
        brand_kpi["recommendation_rate"].tolist(),
        "Brand recommendation rate (%)",
        figures_dir / "brand_recommendation_rate.png",
        x_suffix="%",
    )
    product_labels = [f"{r.brand}: {shorten(r.product_name, 42)}" for r in product_volume.itertuples()]
    draw_bar_chart(
        product_labels,
        product_volume["review_count"].tolist(),
        "Top 5 products per brand - sampled review volume",
        figures_dir / "product_volume.png",
        width=1200,
        height=760,
    )
    skin_plot = skin_type_kpi.sort_values("review_count", ascending=False).head(18)
    skin_labels = [f"{r.brand}: {r.skin_type}" for r in skin_plot.itertuples()]
    draw_bar_chart(
        skin_labels,
        skin_plot["recommendation_rate"].tolist(),
        "Skin type recommendation rate (%)",
        figures_dir / "skin_type_recommendation_rate.png",
        x_suffix="%",
        width=1050,
        height=740,
    )
    draw_line_chart(monthly_trend, "month", "avg_rating", "Monthly average rating", figures_dir / "monthly_avg_rating.png")
    draw_line_chart(
        quarterly_trend,
        "quarter",
        "recommendation_rate",
        "Quarterly recommendation rate (%)",
        figures_dir / "quarterly_recommendation_rate.png",
    )


def font(size: int = 16) -> ImageFont.ImageFont:
    candidates = [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/calibri.ttf",
        "C:/Windows/Fonts/msyh.ttc",
    ]
    for path in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, size=size)
    return ImageFont.load_default()


def shorten(text: object, max_len: int) -> str:
    value = str(text)
    return value if len(value) <= max_len else value[: max_len - 3] + "..."


def palette() -> list[tuple[int, int, int]]:
    return [(45, 108, 172), (237, 133, 75), (78, 153, 93), (170, 92, 148), (110, 110, 110)]


def draw_bar_chart(
    labels: list[object],
    values: list[object],
    title: str,
    path: Path,
    x_suffix: str = "",
    width: int = 920,
    height: int = 520,
) -> None:
    labels = [shorten(label, 58) for label in labels]
    vals = [float(v) if pd.notna(v) else 0.0 for v in values]
    img = Image.new("RGB", (width, height), "white")
    d = ImageDraw.Draw(img)
    title_font = font(24)
    label_font = font(14)
    small_font = font(13)
    left = min(430, max(170, max((text_width(d, label, label_font) for label in labels), default=120) + 28))
    top, right, bottom = 70, 45, 45
    plot_w = width - left - right
    row_h = max(20, (height - top - bottom) / max(1, len(vals)))
    max_val = max(vals) if vals else 1
    max_val = max_val if max_val > 0 else 1

    d.text((24, 22), title, fill=(30, 35, 40), font=title_font)
    colors = palette()
    for i, (label, val) in enumerate(zip(labels, vals)):
        y = top + i * row_h
        bar_h = max(10, row_h * 0.58)
        d.text((24, y + (row_h - 14) / 2 - 2), label, fill=(45, 45, 45), font=label_font)
        d.rectangle((left, y + (row_h - bar_h) / 2, left + plot_w, y + (row_h + bar_h) / 2), fill=(239, 242, 245))
        bar_w = plot_w * (val / max_val)
        d.rectangle((left, y + (row_h - bar_h) / 2, left + bar_w, y + (row_h + bar_h) / 2), fill=colors[i % len(colors)])
        value_label = f"{val:,.1f}{x_suffix}" if x_suffix else f"{val:,.0f}"
        d.text((min(left + bar_w + 6, width - 100), y + (row_h - 14) / 2 - 2), value_label, fill=(30, 35, 40), font=small_font)
    img.save(path)


def draw_line_chart(df: pd.DataFrame, x_col: str, y_col: str, title: str, path: Path) -> None:
    if df.empty:
        return
    width, height = 1120, 560
    img = Image.new("RGB", (width, height), "white")
    d = ImageDraw.Draw(img)
    title_font = font(24)
    label_font = font(13)
    d.text((24, 22), title, fill=(30, 35, 40), font=title_font)
    left, top, right, bottom = 80, 80, 190, 80
    plot_w, plot_h = width - left - right, height - top - bottom
    plot_df = df.copy().sort_values(x_col)
    x_values = sorted(plot_df[x_col].dropna().unique().tolist())
    y_min = float(plot_df[y_col].min())
    y_max = float(plot_df[y_col].max())
    if y_min == y_max:
        y_min -= 0.5
        y_max += 0.5
    x_pos = {x: left + (i / max(1, len(x_values) - 1)) * plot_w for i, x in enumerate(x_values)}
    colors = dict(zip(sorted(plot_df["brand"].unique()), palette()))

    d.rectangle((left, top, left + plot_w, top + plot_h), outline=(210, 215, 220), width=1)
    for j in range(5):
        y = top + j * plot_h / 4
        d.line((left, y, left + plot_w, y), fill=(235, 238, 242))
        val = y_max - j * (y_max - y_min) / 4
        d.text((16, y - 8), f"{val:.1f}", fill=(70, 70, 70), font=label_font)

    for brand, group in plot_df.groupby("brand"):
        pts = []
        for row in group.itertuples():
            x = x_pos[getattr(row, x_col)]
            y_val = float(getattr(row, y_col))
            y = top + plot_h - ((y_val - y_min) / (y_max - y_min)) * plot_h
            pts.append((x, y))
        color = colors[brand]
        if len(pts) >= 2:
            d.line(pts, fill=color, width=3)
        for x, y in pts:
            d.ellipse((x - 4, y - 4, x + 4, y + 4), fill=color)

    step = max(1, len(x_values) // 10)
    for i, label in enumerate(x_values):
        if i % step == 0 or i == len(x_values) - 1:
            x = x_pos[label]
            d.text((x - 28, top + plot_h + 16), str(label), fill=(60, 60, 60), font=label_font)
    legend_x = left + plot_w + 25
    for i, (brand, color) in enumerate(colors.items()):
        y = top + i * 28
        d.rectangle((legend_x, y + 4, legend_x + 18, y + 18), fill=color)
        d.text((legend_x + 26, y), brand, fill=(45, 45, 45), font=label_font)
    img.save(path)


def text_width(draw: ImageDraw.ImageDraw, text: str, draw_font: ImageFont.ImageFont) -> int:
    bbox = draw.textbbox((0, 0), text, font=draw_font)
    return bbox[2] - bbox[0]


def write_brief(
    path: Path,
    raw_rows: int,
    cleaned_rows: int,
    brand_filtered: pd.DataFrame,
    top_filtered: pd.DataFrame,
    sampled: pd.DataFrame,
    table_map: dict[str, pd.DataFrame],
) -> Path:
    brand_kpi = table_map["brand_kpi"]
    product_volume = table_map["product_volume"]
    skin_type_kpi = table_map["skin_type_kpi"]
    monthly_trend = table_map["monthly_trend"]
    quarterly_trend = table_map["quarterly_trend"]
    missing = table_map["missingness_summary"]

    date_min = sampled["submission_date"].min()
    date_max = sampled["submission_date"].max()
    top_missing = (
        missing[missing["dataset"].eq("sampled_target_brands_top_products")]
        .sort_values("missing_pct", ascending=False)
        .head(8)
    )

    lines = [
        "# Task 1 Brief - Sephora Reviews",
        "",
        "## 1. 清洗后数据量与范围",
        f"- 原始合并数据量：{raw_rows:,} 条。",
        f"- 全品牌基础清洗后：{cleaned_rows:,} 条。",
        f"- 三个目标品牌清洗后：{len(brand_filtered):,} 条。",
        f"- 每品牌 top 5 产品筛选后：{len(top_filtered):,} 条。",
        f"- 每品牌最多抽样 2,000 条后的分析样本：{len(sampled):,} 条。",
        f"- 样本时间范围：{date_min:%Y-%m-%d} 至 {date_max:%Y-%m-%d}。" if pd.notna(date_min) else "- 样本没有可解析日期。",
        "",
        "## 2. 缺失值概览",
    ]
    for _, row in top_missing.iterrows():
        lines.append(f"- {row['column']}: {int(row['missing_count']):,} missing ({row['missing_pct']}%).")

    lines.extend(["", "## 3. 品牌 KPI"])
    for _, row in brand_kpi.iterrows():
        lines.append(
            f"- {row['brand']}: {int(row['review_count']):,} reviews, "
            f"avg rating {row['avg_rating']}, recommendation rate {row['recommendation_rate']}%, "
            f"avg price ${row['avg_price_usd']}."
        )

    lines.extend(["", "## 4. 产品评论分布"])
    for brand, group in product_volume.groupby("brand"):
        products = "; ".join(f"{r.product_name} ({int(r.review_count)})" for r in group.itertuples())
        lines.append(f"- {brand}: {products}.")

    lines.extend(["", "## 5. Skin type 分组"])
    for brand, group in skin_type_kpi.sort_values(["brand", "review_count"], ascending=[True, False]).groupby("brand"):
        top_skin = group.head(3)
        desc = "; ".join(
            f"{r.skin_type}: {int(r.review_count)} reviews, rec {r.recommendation_rate}%"
            for r in top_skin.itertuples()
        )
        lines.append(f"- {brand}: {desc}.")

    lines.extend(["", "## 6. 时间稳定性"])
    if not monthly_trend.empty:
        stability = monthly_trend.groupby("brand").agg(
            months=("month", "nunique"),
            avg_rating_sd=("avg_rating", "std"),
            rec_rate_sd=("recommendation_rate", "std"),
            min_reviews=("review_count", "min"),
        ).reset_index()
        for _, row in stability.iterrows():
            lines.append(
                f"- {row['brand']}: 覆盖 {int(row['months'])} 个月，月均评分标准差 "
                f"{row['avg_rating_sd']:.2f}，推荐率标准差 {row['rec_rate_sd']:.1f} 个百分点，"
                f"最小月评论数 {int(row['min_reviews'])}。"
            )
    if not quarterly_trend.empty:
        lines.append("- 季度趋势表已输出，可用于检查低样本月份被聚合后是否仍保持相同方向。")

    path.write_text("\n".join(lines), encoding="utf-8")
    return path
